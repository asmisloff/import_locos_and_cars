from dataclasses import dataclass
import os
from typing import List
from loco import stringValueOf, intValueOf, floatValueOf, \
    nullableStringValueOf, nullableIntValueOf, nullableFloatValueOf, \
    JsonSerializable
import openpyxl as xl
from openpyxl.worksheet import worksheet
from openpyxl import cell

@dataclass
class ResistanceToMotion(JsonSerializable):
    componentRail: List[float | None]
    continuousRail: List[float | None]

@dataclass
class Car:
    name: str
    numberOfAxles: str
    selfWeight: float
    weight: float
    length: float
    resistanceToMotion: ResistanceToMotion

    def __init__(self, wb: xl.Workbook):
        mainParams: worksheet.Worksheet = wb["Основные параметры"]
        self.name = f"*{stringValueOf(mainParams.cell(1, 2))}*"
        self.numberOfAxles = self.__numberOfAxlesFromInt(intValueOf(mainParams.cell(2, 2)))
        self.weight = round(floatValueOf(mainParams.cell(3, 2)), 3)
        self.length = round(floatValueOf(mainParams.cell(4, 2)), 3)

        self.resistanceToMotion = ResistanceToMotion([], [])
        rtm = wb["Осн. удельн. сопр. движ."]
        precisions = [3, 3, 4, 7]
        for colIdx in range(2, 6):
            comp = nullableFloatValueOf(rtm.cell(2, colIdx))
            cont = nullableFloatValueOf(rtm.cell(3, colIdx))
            
            if comp is not None:
                comp = round(9.8 * comp, precisions[colIdx - 2])
            elif colIdx > 2:
                raise Exception("Отсутствует значение коэф. сопр.")
            if cont is not None:
                cont = round(9.8 * cont, precisions[colIdx - 2])
            elif colIdx > 2:
                raise Exception("Отсутствует значение коэф. сопр.")
            
            self.resistanceToMotion.componentRail.append(comp)
            self.resistanceToMotion.continuousRail.append(cont)
        
    def __numberOfAxlesFromInt(self, n) -> str:
        return {
            4: "FOUR_AXLES",
            6: "SIX_AXLES",
            8: "EIGHT_AXLES",
            10: "TEN_AXLES",
            12: "TWELVE_AXLES"
        }[n]

    def toSql(self) -> str:
        n = self.name
        noax = self.numberOfAxles
        w = self.weight
        l = self.length
        return f"('true', now(), '{n}', '{noax}', {w}, {l}, '{self.resistanceToMotion.toJson()}')"


if __name__ == "__main__":
    with open("./cars/cars.sql", 'w', encoding="utf-8") as fh:
        for root, dirs, files in os.walk(os.getcwd() + '/cars'):
            xlsFiles= [f for f in files if f.endswith(".xlsx")]
            cars: List[Car] = []
            for i, f in enumerate(xlsFiles):
                if f.endswith(".xlsx"):
                    wb = xl.load_workbook("./cars/" + f)
                    print(f)
                    car = Car(wb)
                    print(f"{car.name} ({f}) -- успешно\n")
                    cars.append(car)
            values = ",\n".join([car.toSql() for car in cars])
            fh.write("insert into asu_ter_k_main_car (active, change_time, name, number_of_axles, weight, length, resistance_to_motion) values\n")
            fh.write(values)