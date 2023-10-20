from dataclasses import dataclass
import os
import sys
from typing import List
import openpyxl as xl
from openpyxl.worksheet import worksheet
from openpyxl import cell
import json

def stringValueOf(cell: cell.cell.Cell) -> str:
    return str(cell.value).strip()

def intValueOf(cell: cell.cell.Cell) -> int:
    return int(stringValueOf(cell))

def floatValueOf(cell: cell.cell.Cell) -> float:
    return float(stringValueOf(cell))

def nullableStringValueOf(cell: cell.cell.Cell) -> str | None:
    if cell.value is None:
        return None
    else:
        return stringValueOf(cell)

def nullableIntValueOf(cell: cell.cell.Cell) -> int | None:
    if cell.value is None:
        return None
    else:
        return int(stringValueOf(cell))

def nullableFloatValueOf(cell: cell.cell.Cell) -> float | None:
    if cell.value is None:
        return None
    else:
        return float(stringValueOf(cell))
    
class JsonSerializable:
    def toJson(self) -> str:
        return json.dumps(self, default=lambda o: o.__dict__, ensure_ascii=False)

@dataclass
class LocoMainParameters:
    name: str
    type: str
    current: str
    power: float
    weight: float
    length: float
    max_speed: int
    motor_type: str | None
    power_self_consumption: float | None
    amperage_self_consumption: float | None

    def __init__(self, ws: worksheet.Worksheet):
        self.name = "*" + stringValueOf(ws.cell(1, 2)) + "*"
        self.type = stringValueOf(ws.cell(2, 2))
        match self.type:
            case "груз":
                self.type = "FREIGHT_LOCOMOTIVE"
            case "пасс":
                self.type = "PASSENGER_LOCOMOTIVE"
            case "электричка":
                self.type = "ELECTRIC_TRAIN"
            case _:
                raise Exception(f"Неизвестный тип локомотива: '{self.type}'")
        self.current: str = stringValueOf(ws.cell(3, 2))
        match self.current:
            case "3000":
                self.current = "DIRECT_CURRENT"
            case "25000":
                self.current = "ALTERNATING_CURRENT"
            case _:
                raise Exception(f"Неизвестный род тока: {self.current}")
        self.power = nullableFloatValueOf(ws.cell(4, 2)) or 0
        self.weight = floatValueOf(ws.cell(5, 2))
        self.length = floatValueOf(ws.cell(6, 2))
        self.max_speed = intValueOf(ws.cell(7, 2))
        self.motor_type = nullableStringValueOf(ws.cell(8, 2))
        self.amperage_self_consumption = nullableFloatValueOf(ws.cell(10, 2))
        if self.current == "ALTERNATING_CURRENT" or self.amperage_self_consumption is None:
            self.power_self_consumption = nullableFloatValueOf(ws.cell(9, 2))
        else:
            self.power_self_consumption = None
    
    def toString(self) -> str:
        return f"TRUE, now(), '{self.name}', '{self.current}', '{self.type}', {self.power}, {self.weight}, {self.length}, " + \
               f"{self.max_speed}, '{self.motor_type}', {self.power_self_consumption or 'NULL'}, {self.amperage_self_consumption or 'NULL'}"


@dataclass
class ResistanceToMotion(JsonSerializable):
    componentRail: List[float]
    continuousRail: List[float]

@dataclass
class LocomotiveResistanceToMotion(JsonSerializable):
    idleResistanceCoefficients: ResistanceToMotion
    motoringResistanceCoefficients: ResistanceToMotion

    def __init__(self, ws: worksheet.Worksheet):
        self.idleResistanceCoefficients = ResistanceToMotion(
            componentRail=[floatValueOf(ws.cell(3, i)) for i in range(2, 5)],
            continuousRail=[floatValueOf(ws.cell(5, i)) for i in range(2, 5)]
        )
        self.motoringResistanceCoefficients = ResistanceToMotion(
            componentRail=[floatValueOf(ws.cell(2, i)) for i in range(2, 5)],
            continuousRail=[floatValueOf(ws.cell(4, i)) for i in range(2, 5)]
        )
        allArrays = (
            self.idleResistanceCoefficients.componentRail, self.idleResistanceCoefficients.continuousRail,
            self.motoringResistanceCoefficients.componentRail, self.motoringResistanceCoefficients.continuousRail
        )
        precisions = (4, 5, 7)
        for a in allArrays:
            for i, k in enumerate(a):
                a[i] = round(9.8 * a[i], precisions[i])


@dataclass
class AcElectricalCharacteristic(JsonSerializable):
    speed: float
    force: float
    motorAmperage: float
    commutateCurrentAmperage: float
    activeCurrentAmperage: float

    def __post_init__(self):
        self.type: str = "AlternateCharacteristic"
        self.force = round(9.8e-3 * self.force, 4)
        if self.commutateCurrentAmperage < 0:
            self.commutateCurrentAmperage = 0
        if self.activeCurrentAmperage < 0:
            self.activeCurrentAmperage = 0
        if self.motorAmperage < 0:
            self.motorAmperage = 0


@dataclass
class DcElectricalCharacteristic(JsonSerializable):
    speed: float
    force: float
    motorAmperage: float
    activeCurrentAmperage: float

    def __post_init__(self):
        self.type: str = "DirectCharacteristic"
        self.force = round(9.8e-3 * self.force, 4)
        if self.activeCurrentAmperage < 0:
            self.activeCurrentAmperage = 0
        if self.motorAmperage < 0:
            self.motorAmperage = 0


@dataclass 
class ElectricalPosition(JsonSerializable):
    name: str
    characteristics: List[AcElectricalCharacteristic | DcElectricalCharacteristic]

    @classmethod
    def load(cls, ws: worksheet.Worksheet, tp: str) -> List["ElectricalPosition"]:
        rowIdx = 2
        chars: List[AcElectricalCharacteristic | DcElectricalCharacteristic] = []
        res: List[ElectricalPosition] = []
        while True:
            name = nullableStringValueOf(ws.cell(rowIdx, 1))
            if name is None:
                break
            rowIdx += 1
            for _ in range(30):
                if nullableFloatValueOf(ws.cell(rowIdx, 2)) is not None:
                    match tp:
                        case "ac":
                            chars.append(AcElectricalCharacteristic(*[floatValueOf(ws.cell(rowIdx, i)) for i in range(2, 7)]))
                        case "dc":
                            chars.append(DcElectricalCharacteristic(*[floatValueOf(ws.cell(rowIdx, i)) for i in range(2, 6)]))
                rowIdx += 1
            res.append(ElectricalPosition(name, chars[:]))
            chars.clear()
        return res


@dataclass
class BrakingCharacteristics(JsonSerializable):
    limit: List[AcElectricalCharacteristic | DcElectricalCharacteristic]
    max: List[AcElectricalCharacteristic | DcElectricalCharacteristic]

    def __init__(self, ws: worksheet.Worksheet, tp: str):
        rowIdx = 2
        chars: List[List[AcElectricalCharacteristic | DcElectricalCharacteristic]] = []
        while True:
            name = nullableStringValueOf(ws.cell(rowIdx, 1))
            if name is None:
                break
            chars.append([])
            rowIdx += 1
            for _ in range(30):
                if nullableFloatValueOf(ws.cell(rowIdx, 2)) is not None:
                    match tp:
                        case "ac":
                            chars[-1].append(AcElectricalCharacteristic(*[floatValueOf(ws.cell(rowIdx, i)) for i in range(2, 7)]))
                        case "dc":
                            chars[-1].append(DcElectricalCharacteristic(*[floatValueOf(ws.cell(rowIdx, i)) for i in range(2, 6)]))
                rowIdx += 1
        self.limit, self.max = chars


@dataclass
class MotorThermalCharacteristic(JsonSerializable):
    motorAmperage: float
    balancingOverheat: float


@dataclass
class MotorThermalCharacteristics(JsonSerializable):
    overheatTolerance: float
    thermalTimeConstant: float
    characteristics: List[MotorThermalCharacteristic]

    def __init__(self, ws: worksheet.Worksheet):
        self.overheatTolerance = nullableFloatValueOf(ws.cell(1, 2)) or 100
        self.thermalTimeConstant = nullableFloatValueOf(ws.cell(2, 2)) or 20
        self.characteristics = []
        for col in range(2, 19):
            amp = nullableFloatValueOf(ws.cell(4, col))
            if amp is None:
                break
            self.characteristics.append(
                MotorThermalCharacteristic(amp, floatValueOf(ws.cell(5, col)))
            )


@dataclass
class Locomotive(JsonSerializable):
    mainParams: LocoMainParameters
    rtm: LocomotiveResistanceToMotion
    tractiveChars: List[ElectricalPosition]
    recupChars: BrakingCharacteristics
    thermalChars: MotorThermalCharacteristics

    def __init__(self, wb: xl.Workbook):
        self.mainParams = LocoMainParameters(wb["Основные параметры"])
        tp = "dc" if self.mainParams.current == "DIRECT_CURRENT" else "ac"
        self.rtm = LocomotiveResistanceToMotion(wb["Осн. удельн. сопр. движ."])
        self.tractiveChars = ElectricalPosition.load(wb["Хар. тяг. режима"], tp)
        self.recupChars = BrakingCharacteristics(wb["Хар. рекуп. торм."], tp)
        self.thermalChars = MotorThermalCharacteristics(wb["Тепловые хар. двиг."])
    
    def insertQuery(self) -> str:
        name = self.__valueOrNull(self.mainParams.name)
        current = self.__valueOrNull(self.mainParams.current)
        tp = self.__valueOrNull(self.mainParams.type)
        p = self.__valueOrNull(self.mainParams.power)
        weight = self.__valueOrNull(self.mainParams.weight)
        length = self.__valueOrNull(self.mainParams.length)
        speed = self.__valueOrNull(self.mainParams.max_speed)
        mt = self.__valueOrNull(self.mainParams.motor_type)
        psc = self.__valueOrNull(self.mainParams.power_self_consumption)
        asc = self.__valueOrNull(self.mainParams.amperage_self_consumption)
        return f"""
INSERT INTO asu_ter.asu_ter_k_main_locomotive (active, change_time, name, current, type, power, weight, length,
                                max_speed, motor_type, power_self_consumption, amperage_self_consumption,
                                motor_thermal_characteristics, resistance_to_motion, electrical_characteristics,
                                braking_characteristics)
values (
    TRUE, now(), {name}, {current}, {tp}, {p}, {weight}, {length}, {speed}, {mt}, {psc}, {asc}, 
    '{self.thermalChars.toJson()}',
    '{self.rtm.toJson()}',
    '{json.dumps(self.tractiveChars, default=lambda o: o.__dict__, ensure_ascii=False)}',
    '{self.recupChars.toJson()}'
);
""".strip()
    
    def __valueOrNull(self, v: int | float | str | None) -> str:
        if v is None:
            return "NULL"
        elif type(v) == str:
            return f"'{v}'"
        else:
            return str(v)


if __name__ == "__main__":
    with open("./loco/loco.sql", 'w', encoding="utf-8") as fh:
        for root, dirs, files in os.walk(os.getcwd() + '/loco'):
            xlsFiles= [f for f in files if f.endswith(".xlsx")]
            for i, f in enumerate(xlsFiles):
                if f.endswith(".xlsx"):
                    wb = xl.load_workbook("./loco/" + f)
                    print(f)
                    loco = Locomotive(wb)
                    print(f"{loco.mainParams.name} ({f}) -- успешно\n")
                    fh.write(loco.insertQuery())
                    if i != len(xlsFiles) - 1:
                        fh.write("\n\n\n")